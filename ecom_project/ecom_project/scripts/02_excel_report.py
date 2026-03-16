"""
Script 02: Excel Summary Report Generator
Dataset : data/orders_clean.csv
Run from: ecom_project/ folder
Command : python scripts/02_excel_report.py
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_csv("data/orders_clean.csv", parse_dates=["order_date","ship_date"])

wb = Workbook()

# ── LIGHT THEME PALETTE ───────────────────────────────────────────────────────
H_FILL  = PatternFill("solid", start_color="E94560")
H_FONT  = Font(color="FFFFFF", bold=True, size=11)
A_FILL  = PatternFill("solid", start_color="FFFFFF")
A_FONT  = Font(color="1A1A2E", size=10)
B_FILL  = PatternFill("solid", start_color="FFF5F7")
B_FONT  = Font(color="1A1A2E", size=10)
TITLE_FONT = Font(color="E94560", bold=True, size=16)
BORDER  = Border(
    bottom=Side(style="thin", color="E8E8F0"),
    top=Side(style="thin", color="E8E8F0"),
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = H_FILL
        cell.font      = H_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER

def style_row(ws, row, cols, alt=False):
    fill = B_FILL if alt else A_FILL
    font = B_FONT if alt else A_FONT
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER

# ── SHEET 1: KPI SUMMARY ─────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "KPI Summary"
ws1.sheet_properties.tabColor = "E94560"

ws1["B2"].value = "E-COMMERCE ANALYTICS — KPI SUMMARY"
ws1["B2"].font  = TITLE_FONT
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 38
ws1.column_dimensions["C"].width = 28

kpis = [
    ("Total Orders",               f"{df.shape[0]:,}"),
    ("Unique Customers",           df["customer_id"].nunique()),
    ("Unique Products",            df["product_name"].nunique()),
    ("Gross Revenue",              f"₹{df['revenue'].sum():,.2f}"),
    ("Total Discounts Given",      f"₹{df['discount_amt'].sum():,.2f}"),
    ("Net Revenue",                f"₹{df['net_revenue'].sum():,.2f}"),
    ("Avg Order Value",            f"₹{df['net_revenue'].mean():,.2f}"),
    ("Return Rate",                f"{df['is_returned'].mean()*100:.1f}%"),
    ("Avg Customer Rating",        f"{df['rating'].mean():.2f} / 5"),
    ("Avg Shipping Days",          f"{df['ship_days'].mean():.1f} days"),
    ("Best Category",              df.groupby('category')['net_revenue'].sum().idxmax()),
    ("Best Region",                df.groupby('region')['net_revenue'].sum().idxmax()),
    ("Top Customer",               df.groupby('customer_name')['net_revenue'].sum().idxmax()),
    ("Most Used Payment",          df['payment_method'].value_counts().idxmax()),
]

ws1["B4"].value = "Metric"
ws1["C4"].value = "Value"
style_header(ws1, 4, 3)
ws1.row_dimensions[4].height = 22

for i, (metric, value) in enumerate(kpis):
    r = i + 5
    ws1.cell(row=r, column=2, value=metric)
    ws1.cell(row=r, column=3, value=str(value))
    style_row(ws1, r, 3, alt=(i % 2 == 0))
    ws1.row_dimensions[r].height = 20

# ── SHEET 2: CATEGORY ANALYSIS ───────────────────────────────────────────────
ws2 = wb.create_sheet("By Category")
ws2.sheet_properties.tabColor = "4ECDC4"

cat = (
    df.groupby("category")
    .agg(Orders=("order_id","count"), Gross=("revenue","sum"),
         Discounts=("discount_amt","sum"), Net=("net_revenue","sum"),
         Avg_Order=("net_revenue","mean"), Avg_Rating=("rating","mean"),
         Return_Rate=("is_returned","mean"))
    .sort_values("Net", ascending=False).reset_index()
)

headers = ["Category","Orders","Gross Revenue","Discounts","Net Revenue","Avg Order","Avg Rating","Return Rate"]
col_widths = [18,10,16,14,14,12,12,14]
for c,(h,w) in enumerate(zip(headers,col_widths),1):
    ws2.cell(row=1,column=c,value=h)
    ws2.column_dimensions[get_column_letter(c)].width = w
style_header(ws2, 1, len(headers))

for i, row in cat.iterrows():
    r = i + 2
    vals = [row["category"], int(row["Orders"]),
            f"₹{row['Gross']:.2f}", f"₹{row['Discounts']:.2f}",
            f"₹{row['Net']:.2f}", f"₹{row['Avg_Order']:.2f}",
            f"{row['Avg_Rating']:.2f}", f"{row['Return_Rate']*100:.1f}%"]
    for c,v in enumerate(vals,1):
        ws2.cell(row=r,column=c,value=v)
    style_row(ws2, r, len(headers), alt=(i%2==0))

# ── SHEET 3: MONTHLY TREND ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Monthly Trend")
ws3.sheet_properties.tabColor = "F5A623"

monthly = (
    df.groupby(["year","month_num","month"])
    .agg(Orders=("order_id","count"), Net_Revenue=("net_revenue","sum"))
    .reset_index().sort_values(["year","month_num"])
)

headers = ["Year","Month #","Month","Orders","Net Revenue"]
for c,(h,w) in enumerate(zip(headers,[10,10,14,10,16]),1):
    ws3.cell(row=1,column=c,value=h)
    ws3.column_dimensions[get_column_letter(c)].width = w
style_header(ws3, 1, len(headers))

for i,(_, row) in enumerate(monthly.iterrows()):
    r = i + 2
    ws3.cell(row=r,column=1,value=int(row["year"]))
    ws3.cell(row=r,column=2,value=int(row["month_num"]))
    ws3.cell(row=r,column=3,value=row["month"])
    ws3.cell(row=r,column=4,value=int(row["Orders"]))
    ws3.cell(row=r,column=5,value=round(row["Net_Revenue"],2))
    style_row(ws3, r, len(headers), alt=(i%2==0))

# ── SHEET 4: TOP CUSTOMERS ───────────────────────────────────────────────────
ws4 = wb.create_sheet("Top Customers")
ws4.sheet_properties.tabColor = "A78BFA"

customers = (
    df.groupby(["customer_id","customer_name"])
    .agg(Orders=("order_id","count"), LTV=("net_revenue","sum"),
         Avg_Order=("net_revenue","mean"), Avg_Rating=("rating","mean"))
    .sort_values("LTV",ascending=False).reset_index()
)

headers = ["Customer ID","Name","Orders","Lifetime Value","Avg Order","Avg Rating"]
for c,(h,w) in enumerate(zip(headers,[14,20,10,16,12,12]),1):
    ws4.cell(row=1,column=c,value=h)
    ws4.column_dimensions[get_column_letter(c)].width = w
style_header(ws4, 1, len(headers))

for i,row in customers.iterrows():
    r = i + 2
    vals = [row["customer_id"], row["customer_name"], int(row["Orders"]),
            f"₹{row['LTV']:.2f}", f"₹{row['Avg_Order']:.2f}", f"{row['Avg_Rating']:.2f}"]
    for c,v in enumerate(vals,1):
        ws4.cell(row=r,column=c,value=v)
    style_row(ws4, r, len(headers), alt=(i%2==0))

# ── SHEET 5: RAW DATA ────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Raw Data")
ws5.sheet_properties.tabColor = "8A8AAA"

raw_cols = ["order_id","order_date","customer_name","region","category",
            "product_name","quantity","unit_price","discount_pct","net_revenue",
            "payment_method","ship_mode","ship_days","returned","rating"]
raw = df[raw_cols].copy()
raw["order_date"] = raw["order_date"].dt.strftime("%Y-%m-%d")

for c,col in enumerate(raw_cols,1):
    ws5.cell(row=1,column=c,value=col.replace("_"," ").title())
    ws5.column_dimensions[get_column_letter(c)].width = 16
style_header(ws5, 1, len(raw_cols))

for i,row in raw.iterrows():
    r = i + 2
    for c,val in enumerate(row.values,1):
        ws5.cell(row=r,column=c,value=val)
    style_row(ws5, r, len(raw_cols), alt=(i%2==0))

wb.save("outputs/ecommerce_report.xlsx")
print("✓ Excel report saved → outputs/ecommerce_report.xlsx")
print("  Sheets: KPI Summary | By Category | Monthly Trend | Top Customers | Raw Data")
